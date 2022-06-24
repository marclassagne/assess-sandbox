from openpyxl import load_workbook
import os
import json
import traceback
import sys
import fit


def convert_utility(s):
    """ Convert utility strings in excel to adapted format """
    if s == 'quadratic':
        return('quad')
    elif s == 'expo-power':
        return(s)
    else:
        return(s[:3])


def importation(file):

    try:
        wb = load_workbook(filename=file, read_only=True)

        mySession = {'attributes': [], 'k_calculus': [{'method': 'multiplicative', 'active': True, 'k': [], 'GK':None, 'GU':None}, {'method': 'multilinear', 'active': False, 'k': [], 'GK':None, 'GU':None}], "settings": {
            "decimals_equations": 3,
            "decimals_dpl": 8,
            "proba_ce": 0.3,
            "proba_le": 0.3,
            "language": "english",
            "display": "trees"
        }}

        for sheet in wb:
            myAttribut = {}
            ws = wb[sheet.title]  # ws is now an IterableWorksheet
            if sheet.title == "Multi attribute multilinear" or sheet.title == "Multi attribute multiplicative":
                continue

            myAttribut['name'] = ws['B2'].value
            myAttribut['type'] = ws['B3'].value
            myAttribut['unit'] = ws['B4'].value
            myAttribut['method'] = ws['B5'].value
            myAttribut['mode'] = ws['B6'].value
            myAttribut['checked'] = ws['B7'].value
            myAttribut['completed'] = ws['B8'].value

            myAttribut['val_min'] = ws['E2'].value
            myAttribut['val_med'] = []
            mesPoints = {}
            number = 0

            i = 3
            while ws['D' + str(i)].value[0] == 'I':
                myAttribut['val_med'].append(str(ws['E' + str(i)].value))
                if ws['F' + str(i)].value != None:
                    mesPoints[ws['E'+str(i)].value] = ws['F'+str(i)].value
                    number += 1
                i += 1

            myAttribut['val_max'] = ws['E' + str(i)].value

            myAttribut['questionnaire'] = {}

            myAttribut['questionnaire']['points'] = mesPoints
            myAttribut['questionnaire']['number'] = number

            myAttribut['utility'] = {}
            i = 2
            while ws['I' + str(i)].value != None:
                dic = {}
                dic['a'] = ws['I' + str(i+1)].value
                dic['b'] = ws['I' + str(i+2)].value
                dic['c'] = ws['I' + str(i+3)].value
                dic['d'] = ws['I' + str(i+4)].value
                dic['r2'] = ws['I' + str(i+5)].value
                myAttribut['utility'][convert_utility(
                    ws['I' + str(i)].value)] = dic
                i += 15

            mySession['attributes'].append(myAttribut)

        for sheet in wb:
            ws = wb[sheet.title]  # ws is now an IterableWorksheet
            if sheet.title != "Multi attribute multilinear" and sheet.title != "Multi attribute multiplicative":
                continue

            ligne = 2
            mesK = []

            while ws['C'+str(ligne)].value != None:

                mesK.append({'ID': ws['A'+str(ligne)].value, 'ID_attribute': json.loads(ws['D'+str(ligne)].value),
                            'attribute': json.loads(ws['C'+str(ligne)].value), 'value': ws['B'+str(ligne)].value})
                ligne = ligne+1

            if sheet.title == "Multi attribute multilinear":
                mySession['k_calculus'][1]['k'] = mesK
                mySession['k_calculus'][1]['GK'] = ws['B'+str(ligne)].value

            elif sheet.title == "Multi attribute multiplicative":
                mySession['k_calculus'][0]['k'] = mesK
                mySession['k_calculus'][0]['GK'] = ws['B'+str(ligne)].value

            try:
                GU = {'U': None, 'utilities': [], 'k': mesK}

                GU['U'] = ws['B'+str(len(mesK)+5)].value

                ligne = 2
                while ws['E'+str(ligne)].value != None:
                    utilityType = ws['E'+str(ligne)].value
                    if utilityType == None:  # we brake all
                        ligne = 2
                        break
                    ID_attribute = ws['D'+str(ligne)
                                      ].value.replace("[", "").replace("]", "")
                    monAttribut = mySession['attributes'][int(ID_attribute)]

                    points = monAttribut['questionnaire']['points'][:]
                    if monAttribut['mode'] == "normal":
                        points.append([monAttribut['val_max'], 1])
                        points.append([monAttribut['val_min'], 0])
                    else:
                        points.append([monAttribut['val_max'], 0])
                        points.append([monAttribut['val_min'], 1])

                    allUtilities = fit.regressions(points, True)
                    for myUtility in allUtilities:
                        if myUtility['type'] == utilityType:
                            GU['utilities'].append(myUtility)
                            break

                    ligne = ligne+1

                if ligne != 2:  # If we have some utilityType function, ligne is different from 2
                    if sheet.title == "Multi attribute multilinear":
                        mySession['k_calculus'][1]['GU'] = GU

                    elif sheet.title == "Multi attribute multiplicative":
                        mySession['k_calculus'][0]['GU'] = GU

            except:  # if it doesn't worl because there is no U or utilities type
                pass

        os.remove(file)
        return {'success': True, 'data': mySession}
    except (Exception, err):
        os.remove(file)
        return {'success': False, 'data': traceback.format_exc()}
